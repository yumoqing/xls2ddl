import os
import sys
from traceback import print_exc
from openpyxl import load_workbook
from appPublic.myjson import loadf,dumpf,dumps
from appPublic.dictObject import DictObject
	
class TypeConvert:
	def conv(self,typ,v):
		if typ is None:
			return v
		f = getattr(self,'to_'+typ,None)
		if f is None:
			return v
		return f(v)
	
	def to_int(self,v):
		try:
			return int(v)
		except:
			return 0
	
	def to_float(self,v):
		try:
			return float(v)
		except:
			return 0.0
	
	def to_str(self,v):
		try:
			return str(v)
		except:
			return ''
	def to_json(self,v):
		if v == '':
			return v
		try:
			return loads(v)
		except:
			return v
			
	def to_date(self,v):
		return v
	
	def to_time(self,v):
		return v
	
	def to_timestamp(self,v):
		return v

	def to_cruddata(self,v):
		vs = v.split('"',3)
		if vs < 3:
			return v
		fn = vs[1]
		d = CRUDData(fn)
		try:
			data = d.get_data()
		except Exception as e:
			print_exc()
			print(e)
			return v
		if vs[2] is None:
			return data
		cmd = "d%s" % vs[2]
		ret=eval(cmd,{'d':data})
		return ret
	
	def to_xlsxdata(self,v):
		vs = v.split('"',3)
		if vs < 3:
			return v
		fn = vs[1]
		d = XLSXData(fn)
		try:
			data = d.get_data()
		except Exception as e:
			print_exc()
			print(e)
			return v
		if vs[2] is None:
			return data
		cmd = "d%s" % vs[2]
		ret=eval(cmd,{'d':data})
		return ret
	
class CRUDException(Exception):
	def __init__(self,xlsfile,errmsg,*args,**argv):
		Exception.__init__(self,*args,**argv)
		self.xlsfile = xlsfile
		self.errmsg = errmsg
	
	def __str__(self):
		return 'filename:' + self.xlsfile+' error:' + self.errmsg

class XLSXData(object):
	def __init__(self,xlsxfile, book=None):
		self.xlsxfile = xlsxfile
		if book is None:
			self.book = load_workbook(filename=xlsxfile)
		else:
			self.book = book
		self._read()

	def get_data(self):
		return self.data

	def readRecords(self,name,sheet):
		i = 1
		recs = []
		fields = []
		tc = TypeConvert()

		for i,row in enumerate(sheet.values):
			if i==0:
				fields = self.getFieldNames(row)
				continue
			rec = {}
			for j, a in enumerate(row):
				if a is None:
					continue
				k = fields[j][0]
				v = tc.conv(fields[j][1],a)
				rec[k] = v
			if rec == {}:
				continue
			o = DictObject(**rec)
			recs.append(o)
		return {name:recs}
	 
	def _read(self):
		ret = {}
		for i,s in enumerate(self.book.worksheets):
			ret.update(self.readRecords(self.book.sheetnames[i], s))
		self.data = DictObject(**ret)

	def getFieldNames(self,row):
		fs = []
		for i,f in enumerate(row):
			if f is None:
				f = 'F_' + str(i)
			else:
				if type(f) != type(""):
					f = 'F_' + str(f)
				"""
				else:
					f = f.encode('utf-8')
				"""
			b=f.split(':')
			if len(b) < 2:
				b.append(None)
			fs.append(b)
			i+= 1
		return fs
	  
class CRUDData(XLSXData):
	@classmethod
	def isMe(self,book):
		if book is None:
			return False

		names = book.sheetnames
		if 'summary' not in names:
			return False
		if 'fields' not in names:
			return False
		if 'validation' not in names:
			return False
		return True
	
	def _read(self):
		super()._read()
		d = self.data
		if not 'summary' in d.keys():
			raise CRUDException(self.xlsxfile,'summary sheet missing')
		if not 'fields' in d.keys():
			raise CRUDException(self.xlsxfile,'fields sheet missing')
		if not 'validation' in d.keys():
			raise CRUDException(self.xlsxfile,'validation sheet missing')
		if len(d['summary']) != 1:
			raise CRUDException(self.xlsxfile, 'Not summary or more than one summary')
		self.convPrimary()
		self.convForeignkey()
		self.convIndex()
		self.check_codes_fields()
  
	def convPrimary(self):
		d = self.data
		v = d['summary'][0]['primary']
		v = v.split(',')
		self.data['summary'][0]['primary'] = v
		self.check_primary_fields()
		
	def check_primary_fields(self):
		primarys = self.data['summary'][0]['primary']
		if primarys is None:
			raise CRUDException(self.xlsxfile, 'primary is None')
		for p in primarys:
			r = self.check_field(p)
			if not r:
				raise CRUDException(
					self.xlsxfile, f'primary error({p})')
	
	def check_codes_fields(self):
		if 'codes' not in self.data.keys():
			return
		for f in self.data['codes']:
			r = self.check_field(f['field'])
			if not r:
				raise CRUDException(
					self.xlsxfile, 
					f'code definintion error({f["field"]})')


	def check_field(self, fieldname):
		return fieldname in [f['name'] for f in self.data['fields']]

	def convForeignkey(self):
		data = self.data
		vs = data['validation']
		nvs = []
		for v in vs:
			if v['oper'] == 'fk':
				m = v['value']
				des= m.split(':')
				if len(des) != 3:
					raise CRUDException(self.xlsxfile,'fk value error:%s' % m)
				v['value'] = {'table':des[0],'value':des[1],'title':des[2]}
			nvs.append(v)
		data['validation'] = nvs
		self.data = data
 
	def getFieldByNmae(self,fields,name):
		for f in fields:
			if f['name'] == name:
				return f

	def getFKs(self,validation):
		fks = []
		for v in validation:
			if v['oepr'] == 'fk':
				fks.append(v)
		return fks

	def getIDXs(self,validation):
		idxs = []
		for v in validation:
			if v['oper'] == 'idx':
				idxs.append(v)
		return idxs

	def convIndex(self):
		data = self.data
		vs = data['validation']
		nvs = []
		for v in vs:
			if v['oper'] == 'idx':
				idx = {}
				idx['name'] = v['name']
				m = v['value']
				des= m.split(':')
				if len(des) != 2:
					raise CRUDException(self.xlsxfile,'idx value format:idx_type:keylist:%s' % m)
				idx['idxtype'] = des[0]
				idx['idxfields'] = des[1].split(',')
				self.check_field(f for f in idx['idxfields'])
				nvs.append(idx)
		data['indexes'] = nvs
		self.data = data

def xlsxFactory(xlsxfilename):
	def findSubclass(name,klass):
		for k in klass.__subclasses__():
			if k.isMe(name):
				return k
			k1 = findSubclass(name,k)
			if k1 is not None:
				return k1
			return None
	try:
		book = load_workbook(filename=xlsxfilename)
		if book is None:
			print(f'{xlsxfilename} read error')
			return None
		k = findSubclass(book, XLSXData)
		if k is not None:
			xlsx = k(xlsxfilename, book=book)
			return xlsx
		return XLSXData(xlsxfilename, book=book)

	except Exception as e:
		print_exc()
		print(xlsxfilename, 'new class failed\n%s' % str(e))
		print_exc()
		return None

def ValueConvert(s):
	if s[:9] == 'xlsfile::':
		d = xlsxFactory(s[9:])
		return d.get_data()
	if s[:10] == 'jsonfile::':
		return loadf(s[10:])
	return s
	
def paramentHandle(ns):
	for k,v in ns.items():
		ns[k] = ValueConvert(v)
	return ns

if __name__ == '__main__':
	retData = {}
	ns = {}
	datafiles = []
	for a in sys.argv[1:]:
		m = a.split('=',1)
		if len(m)>1:
			ns[m[0]] = m[1]
		else:
			datafiles.append(a)
	
	ns = paramentHandle(ns)
	for f in datafiles:
		ext = os.path.splitext(f)[-1]
		if ext in ['.xlsx','.xls' ]:
			d = xlsxFactory(f)
			data = d.get_data()
			retData.update(data)
	retData.update(ns)
	print( dumps(retData))
	 
